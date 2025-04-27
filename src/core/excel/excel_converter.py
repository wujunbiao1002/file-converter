#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Junly文件转换工具 - Excel转换核心模块
版权所有 (c) 2025 Junly
"""

import os
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
import logging

class ExcelConverter:
    """Excel文档转换类"""
    
    def __init__(self):
        """初始化Excel转换器"""
        self.logger = logging.getLogger(__name__)
    
    def read_excel(self, file_path):
        """读取Excel文档
        
        尝试多种方法读取Excel文件，增强兼容性
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            openpyxl.Workbook: Excel工作簿对象
            
        Raises:
            Exception: 如果所有读取方法都失败时抛出异常
        """
        error_messages = []
        
        # 方法1: 使用openpyxl直接读取
        try:
            self.logger.info(f"尝试使用openpyxl直接读取: {file_path}")
            return openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            error_msg = f"openpyxl直接读取失败: {str(e)}"
            self.logger.warning(error_msg)
            error_messages.append(error_msg)
            
            # 特殊处理：如果错误消息包含"wildcard"，可能是文件格式问题
            if "wildcard" in str(e).lower():
                self.logger.warning("检测到wildcard错误，可能是Excel文件中有通配符或特殊字符")
        
        # 方法2: 使用openpyxl的read_only模式并设置use_iterators=True读取
        try:
            self.logger.info(f"尝试使用openpyxl read_only模式读取: {file_path}")
            # 增加ignore_workbook_corruption=True参数尝试修复损坏的文件
            return openpyxl.load_workbook(
                file_path, 
                data_only=True, 
                read_only=True,
                keep_links=False  # 禁用链接以避免某些错误
            )
        except Exception as e:
            error_msg = f"openpyxl read_only模式读取失败: {str(e)}"
            self.logger.warning(error_msg)
            error_messages.append(error_msg)
        
        # 方法3: 使用pandas读取，然后转换为openpyxl格式
        try:
            self.logger.info(f"尝试使用pandas读取: {file_path}")
            # 添加引擎参数，尝试不同的引擎
            try:
                # 先尝试默认引擎
                excel_file = pd.ExcelFile(file_path)
            except Exception as pd_e:
                self.logger.warning(f"pandas默认引擎失败: {str(pd_e)}，尝试openpyxl引擎")
                # 如果默认引擎失败，尝试openpyxl引擎
                excel_file = pd.ExcelFile(file_path, engine='openpyxl')
            
            sheet_names = excel_file.sheet_names
            
            # 创建新的openpyxl工作簿
            workbook = openpyxl.Workbook()
            # 删除默认创建的工作表
            if "Sheet" in workbook.sheetnames:
                del workbook["Sheet"]
            
            # 处理每个工作表
            for sheet_name in sheet_names:
                # 读取工作表，添加错误处理
                try:
                    df = excel_file.parse(sheet_name)
                    
                    # 创建新工作表
                    sheet = workbook.create_sheet(title=sheet_name)
                    
                    # 写入表头
                    for col_idx, column in enumerate(df.columns, start=1):
                        try:
                            sheet.cell(row=1, column=col_idx, value=str(column))
                        except Exception as col_e:
                            self.logger.warning(f"写入列名时出错: {str(col_e)}")
                            sheet.cell(row=1, column=col_idx, value="[列名错误]")
                    
                    # 写入数据
                    for row_idx, row in df.iterrows():
                        for col_idx, value in enumerate(row, start=1):
                            try:
                                if pd.isna(value):  # 使用pandas的NA检查
                                    sheet.cell(row=row_idx+2, column=col_idx, value="")
                                else:
                                    sheet.cell(row=row_idx+2, column=col_idx, value=str(value))
                            except Exception as val_e:
                                self.logger.warning(f"写入单元格值时出错: {str(val_e)}")
                                sheet.cell(row=row_idx+2, column=col_idx, value="[值错误]")
                except Exception as sheet_e:
                    self.logger.warning(f"处理工作表 {sheet_name} 失败: {str(sheet_e)}")
                    # 创建一个错误信息工作表
                    sheet = workbook.create_sheet(title=f"{sheet_name}(错误)")
                    sheet.cell(row=1, column=1, value=f"无法读取此工作表: {str(sheet_e)}")
            
            return workbook
        except Exception as e:
            error_msg = f"pandas读取失败: {str(e)}"
            self.logger.warning(error_msg)
            error_messages.append(error_msg)
        
        # 方法4: 尝试使用xlrd库读取(对旧版Excel文件更友好)
        try:
            import xlrd
            self.logger.info(f"尝试使用xlrd读取: {file_path}")
            
            # 使用xlrd打开Excel文件
            xls_book = xlrd.open_workbook(file_path)
            
            # 创建新的openpyxl工作簿
            workbook = openpyxl.Workbook()
            if "Sheet" in workbook.sheetnames:
                del workbook["Sheet"]
            
            # 处理每个工作表
            for sheet_name in xls_book.sheet_names():
                xls_sheet = xls_book.sheet_by_name(sheet_name)
                sheet = workbook.create_sheet(title=sheet_name)
                
                # 复制数据
                for row_idx in range(xls_sheet.nrows):
                    for col_idx in range(xls_sheet.ncols):
                        try:
                            value = xls_sheet.cell_value(row_idx, col_idx)
                            if value is None or value == '':
                                sheet.cell(row=row_idx+1, column=col_idx+1, value='')
                            else:
                                sheet.cell(row=row_idx+1, column=col_idx+1, value=str(value))
                        except Exception as cell_e:
                            self.logger.warning(f"处理xlrd单元格值时出错: {str(cell_e)}")
                            sheet.cell(row=row_idx+1, column=col_idx+1, value='[转换错误]')
            
            return workbook
        except ImportError:
            error_msg = "xlrd库未安装，跳过此方法"
            self.logger.warning(error_msg)
            error_messages.append(error_msg)
        except Exception as e:
            error_msg = f"xlrd读取失败: {str(e)}"
            self.logger.warning(error_msg)
            error_messages.append(error_msg)
        
        # 所有读取方法都失败
        raise Exception(f"读取Excel文档失败: 尝试了多种方法但均失败:\n" + "\n".join(error_messages))
    
    def save_as_txt(self, workbook, output_path, delimiter='\t', progress_callback=None):
        """将Excel文档保存为TXT格式"""
        try:
            sheet_count = len(workbook.sheetnames)
            processed_sheets = 0
            
            with open(output_path, 'w', encoding='utf-8') as f:
                for sheet_name in workbook.sheetnames:
                    try:
                        sheet = workbook[sheet_name]
                        
                        # 写入工作表名称作为标题
                        f.write(f"# {sheet_name}\n\n")
                        
                        # 处理特殊情况:如果这是一个只读工作表(read_only=True)
                        if getattr(sheet, '_read_only', False):
                            self.logger.info(f"工作表 {sheet_name} 是只读模式，使用特殊处理")
                            # 获取只读工作表数据
                            rows = self._get_read_only_sheet_data(sheet)
                            for row_data in rows:
                                f.write(delimiter.join(row_data) + '\n')
                        else:
                            # 处理所有行
                            try:
                                for row in sheet.iter_rows():
                                    # 安全地获取并处理单元格值
                                    row_data = []
                                    for cell in row:
                                        try:
                                            if cell.value is None:
                                                row_data.append('')
                                            else:
                                                row_data.append(str(cell.value))
                                        except Exception as e:
                                            self.logger.warning(f"转换单元格值时出错: {str(e)}")
                                            row_data.append('[转换错误]')
                                    
                                    f.write(delimiter.join(row_data) + '\n')
                            except Exception as e:
                                self.logger.warning(f"处理工作表 {sheet_name} 行时出错: {str(e)}")
                                f.write("数据处理错误：无法读取此工作表内容\n")
                        
                        # 工作表之间添加空行
                        f.write('\n\n')
                    except Exception as sheet_e:
                        self.logger.error(f"处理工作表 {sheet_name} 时出错: {str(sheet_e)}")
                        f.write(f"# {sheet_name}\n\n")
                        f.write(f"处理此工作表时出错: {str(sheet_e)}\n\n")
                    
                    # 更新进度
                    processed_sheets += 1
                    if progress_callback:
                        progress_callback(int(processed_sheets / sheet_count * 100))
            
            return True
            
        except Exception as e:
            self.logger.error(f"保存为TXT失败: {str(e)}", exc_info=True)
            raise Exception(f"保存为TXT失败: {str(e)}")
    
    def save_as_markdown(self, workbook, output_path, progress_callback=None):
        """将Excel文档保存为Markdown格式
        
        如果工作簿有多个工作表，将为每个工作表生成单独的MD文件，文件名格式为"原文件名-Sheet名称.md"
        """
        try:
            sheet_count = len(workbook.sheetnames)
            processed_sheets = 0
            
            # 检查是否有多个工作表
            if sheet_count > 1:
                # 将为每个工作表生成单独的文件
                base_name, ext = os.path.splitext(output_path)
                
                for sheet_name in workbook.sheetnames:
                    # 为每个Sheet创建单独的文件名
                    sheet_output_path = f"{base_name}-{sheet_name}.md"
                    
                    try:
                        sheet = workbook[sheet_name]
                        
                        with open(sheet_output_path, 'w', encoding='utf-8') as f:
                            # 处理特殊情况:如果这是一个只读工作表(read_only=True)
                            if getattr(sheet, '_read_only', False):
                                self.logger.info(f"工作表 {sheet_name} 是只读模式，使用特殊处理")
                                # 创建一个简单表格
                                rows = self._get_read_only_sheet_data(sheet)
                            else:
                                # 获取所有行的数据，同时安全地处理所有单元格值
                                rows = []
                                try:
                                    for row in sheet.iter_rows():
                                        row_values = []
                                        for cell in row:
                                            # 安全地获取单元格值，处理所有可能的类型
                                            try:
                                                if cell.value is None:
                                                    row_values.append('')
                                                else:
                                                    row_values.append(str(cell.value))
                                            except Exception as e:
                                                self.logger.warning(f"转换单元格值时出错: {str(e)}")
                                                row_values.append('[转换错误]')
                                        rows.append(row_values)
                                except Exception as e:
                                    self.logger.warning(f"处理工作表 {sheet_name} 行时出错: {str(e)}")
                                    rows = [["数据处理错误：无法读取此工作表内容"]]
                            
                            if not rows:
                                f.write("*空工作表*\n\n")
                                continue
                            
                            # 确定每列的最大宽度（用于对齐）
                            col_widths = []
                            max_cols = max(len(row) for row in rows) if rows else 0
                            
                            for col_idx in range(max_cols):
                                max_width = 0
                                for row in rows:
                                    if col_idx < len(row):
                                        cell_str = row[col_idx]
                                        max_width = max(max_width, len(cell_str))
                                col_widths.append(max_width)
                            
                            # 如果没有列，继续下一个工作表
                            if not col_widths:
                                f.write("*空工作表*\n\n")
                                continue
                            
                            # 第一行作为表头
                            header = rows[0] if rows else []
                            header_row = "| "
                            for col_idx, cell in enumerate(header):
                                if col_idx < len(col_widths):
                                    header_row += cell.ljust(col_widths[col_idx]) + " | "
                            f.write(header_row + "\n")
                            
                            # 分隔行
                            separator = "| "
                            for width in col_widths:
                                separator += "-" * width + " | "
                            f.write(separator + "\n")
                            
                            # 数据行
                            for row_idx, row in enumerate(rows[1:], 1):
                                row_str = "| "
                                for col_idx, cell in enumerate(row):
                                    if col_idx < len(col_widths):
                                        row_str += cell.ljust(col_widths[col_idx]) + " | "
                                f.write(row_str + "\n")
                    
                    except Exception as sheet_e:
                        self.logger.error(f"处理工作表 {sheet_name} 时出错: {str(sheet_e)}")
                        with open(sheet_output_path, 'w', encoding='utf-8') as f:
                            f.write(f"*处理此工作表时出错: {str(sheet_e)}*\n\n")
                    
                    # 更新进度
                    processed_sheets += 1
                    if progress_callback:
                        progress_callback(int(processed_sheets / sheet_count * 100))
                
                return True
            
            else:
                # 单个工作表，使用原始方法处理
                with open(output_path, 'w', encoding='utf-8') as f:
                    sheet_name = workbook.sheetnames[0]
                    try:
                        sheet = workbook[sheet_name]
                        
                        # 处理特殊情况:如果这是一个只读工作表(read_only=True)
                        if getattr(sheet, '_read_only', False):
                            self.logger.info(f"工作表 {sheet_name} 是只读模式，使用特殊处理")
                            # 创建一个简单表格
                            rows = self._get_read_only_sheet_data(sheet)
                        else:
                            # 获取所有行的数据，同时安全地处理所有单元格值
                            rows = []
                            try:
                                for row in sheet.iter_rows():
                                    row_values = []
                                    for cell in row:
                                        # 安全地获取单元格值，处理所有可能的类型
                                        try:
                                            if cell.value is None:
                                                row_values.append('')
                                            else:
                                                row_values.append(str(cell.value))
                                        except Exception as e:
                                            self.logger.warning(f"转换单元格值时出错: {str(e)}")
                                            row_values.append('[转换错误]')
                                    rows.append(row_values)
                            except Exception as e:
                                self.logger.warning(f"处理工作表 {sheet_name} 行时出错: {str(e)}")
                                rows = [["数据处理错误：无法读取此工作表内容"]]
                        
                        if not rows:
                            f.write("*空工作表*\n\n")
                            return True
                        
                        # 确定每列的最大宽度（用于对齐）
                        col_widths = []
                        max_cols = max(len(row) for row in rows) if rows else 0
                        
                        for col_idx in range(max_cols):
                            max_width = 0
                            for row in rows:
                                if col_idx < len(row):
                                    cell_str = row[col_idx]
                                    max_width = max(max_width, len(cell_str))
                            col_widths.append(max_width)
                        
                        # 如果没有列，继续下一个工作表
                        if not col_widths:
                            f.write("*空工作表*\n\n")
                            return True
                        
                        # 第一行作为表头
                        header = rows[0] if rows else []
                        header_row = "| "
                        for col_idx, cell in enumerate(header):
                            if col_idx < len(col_widths):
                                header_row += cell.ljust(col_widths[col_idx]) + " | "
                        f.write(header_row + "\n")
                        
                        # 分隔行
                        separator = "| "
                        for width in col_widths:
                            separator += "-" * width + " | "
                        f.write(separator + "\n")
                        
                        # 数据行
                        for row_idx, row in enumerate(rows[1:], 1):
                            row_str = "| "
                            for col_idx, cell in enumerate(row):
                                if col_idx < len(col_widths):
                                    row_str += cell.ljust(col_widths[col_idx]) + " | "
                            f.write(row_str + "\n")
                        
                    except Exception as sheet_e:
                        self.logger.error(f"处理工作表 {sheet_name} 时出错: {str(sheet_e)}")
                        f.write(f"*处理此工作表时出错: {str(sheet_e)}*\n\n")
                        
                    # 更新进度
                    if progress_callback:
                        progress_callback(100)
                
                return True
            
        except Exception as e:
            self.logger.error(f"保存为Markdown失败: {str(e)}", exc_info=True)
            raise Exception(f"保存为Markdown失败: {str(e)}")
    
    def _get_read_only_sheet_data(self, sheet):
        """专门处理只读工作表的数据"""
        rows = []
        try:
            # 尝试使用.rows属性获取只读工作表数据
            for row in sheet.rows:
                row_values = []
                for cell in row:
                    try:
                        if cell.value is None:
                            row_values.append('')
                        else:
                            row_values.append(str(cell.value))
                    except Exception as e:
                        self.logger.warning(f"转换只读工作表单元格值时出错: {str(e)}")
                        row_values.append('[转换错误]')
                rows.append(row_values)
        except Exception as e:
            self.logger.warning(f"获取只读工作表数据出错，尝试备选方法: {str(e)}")
            # 如果上面的方法失败，使用values属性直接获取值
            try:
                for row in sheet.values:
                    row_values = []
                    for value in row:
                        try:
                            if value is None:
                                row_values.append('')
                            else:
                                row_values.append(str(value))
                        except Exception as val_e:
                            self.logger.warning(f"转换只读工作表值时出错: {str(val_e)}")
                            row_values.append('[转换错误]')
                    rows.append(row_values)
            except Exception as val_err:
                self.logger.error(f"无法获取只读工作表数据: {str(val_err)}")
                rows = [["[无法读取工作表数据]"]]
        
        return rows
    
    def get_sheet_data(self, sheet):
        """获取工作表数据"""
        # 处理特殊情况:如果这是一个只读工作表(read_only=True)
        if getattr(sheet, '_read_only', False):
            self.logger.info(f"工作表是只读模式，使用特殊处理")
            return self._get_read_only_sheet_data(sheet)
        
        # 普通工作表处理
        data = []
        try:
            for row in sheet.iter_rows():
                row_data = []
                for cell in row:
                    try:
                        if cell.value is None:
                            row_data.append('')
                        else:
                            row_data.append(str(cell.value))
                    except Exception as e:
                        self.logger.warning(f"获取单元格值时出错: {str(e)}")
                        row_data.append('[转换错误]')
                data.append(row_data)
        except Exception as e:
            self.logger.error(f"获取工作表数据时出错: {str(e)}")
            data = [["[数据处理错误：无法读取此工作表内容]"]]
        
        return data
    
    def is_empty_row(self, row):
        """检查行是否为空"""
        return all(cell == '' or cell is None for cell in row) 